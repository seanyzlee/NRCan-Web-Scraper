"""
NRCan Article Scraper — core engine
====================================
Can be used as a library (imported by app.py) or as a CLI.

Library usage:
    from scraper import run
    articles = run(sources, keywords, days, run_id, use_dedup, progress_callback)

CLI usage:
    python scraper.py [--days N] [--keywords "kw1 kw2"] [--no-dedup]
"""

import logging
import re
import time
from datetime import datetime, timedelta, timezone
from typing import Callable, Optional
from urllib.parse import urlparse

import feedparser
import requests
from bs4 import BeautifulSoup
from dateutil import parser as dateparser

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# HTTP session
# ---------------------------------------------------------------------------
SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-CA,en;q=0.9",
})
REQUEST_TIMEOUT = 20
REQUEST_DELAY   = 1.0


# ===========================================================================
# Keyword matching
# ===========================================================================

def build_pattern(keywords: list[str]) -> re.Pattern:
    escaped = [re.escape(kw) for kw in keywords]
    return re.compile("|".join(escaped), re.IGNORECASE)


def matches(pattern: re.Pattern, *texts: str) -> bool:
    combined = " ".join(t for t in texts if t)
    return bool(pattern.search(combined))


# ===========================================================================
# Date utilities
# ===========================================================================

def parse_date(raw) -> Optional[datetime]:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        dt = raw
    elif isinstance(raw, time.struct_time):
        dt = datetime(*raw[:6], tzinfo=timezone.utc)
    else:
        try:
            dt = dateparser.parse(str(raw))
        except Exception:
            return None
    if dt and dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def within_window(dt: Optional[datetime], cutoff: datetime) -> bool:
    return dt is None or dt >= cutoff


# ===========================================================================
# Row builder
# ===========================================================================

def _build_row(source: dict, title: str, url: str,
               pub_dt: Optional[datetime], summary: str) -> dict:
    return {
        "Source":     source["name"],
        "Category":   source.get("category", ""),
        "Title":      title.strip(),
        "URL":        url.strip(),
        "Published":  pub_dt.strftime("%Y-%m-%d") if pub_dt else "",
        "Summary":    _clean_html(summary)[:500],
        "Scraped At": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
    }


def _clean_html(text: str) -> str:
    if not text:
        return ""
    return BeautifulSoup(text, "lxml").get_text(separator=" ", strip=True)


# ===========================================================================
# RSS fetcher
# ===========================================================================

def fetch_rss(source: dict, cutoff: datetime, pattern: re.Pattern,
              use_dedup: bool, seen_check: Callable, seen_mark: Callable) -> list[dict]:
    results = []
    for url in source["urls"]:
        log.info("  RSS  %s → %s", source["name"], url)
        try:
            feed = feedparser.parse(url, agent=SESSION.headers["User-Agent"])
        except Exception as exc:
            log.warning("    feedparser error: %s", exc)
            continue

        for entry in feed.entries:
            article_url = entry.get("link", "")
            if not article_url:
                continue
            if use_dedup and seen_check(article_url):
                continue

            title   = entry.get("title", "")
            summary = entry.get("summary", "") or entry.get("description", "")
            pub_dt  = parse_date(entry.get("published_parsed") or entry.get("updated_parsed"))

            if not within_window(pub_dt, cutoff):
                continue
            if not matches(pattern, title, summary):
                continue

            results.append(_build_row(source, title, article_url, pub_dt, summary))
            if use_dedup:
                seen_mark(article_url)

        time.sleep(REQUEST_DELAY)
    return results


# ===========================================================================
# HTML scrapers
# ===========================================================================

def _get_soup(url: str) -> Optional[BeautifulSoup]:
    try:
        resp = SESSION.get(url, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        return BeautifulSoup(resp.text, "lxml")
    except Exception as exc:
        log.warning("    GET failed (%s): %s", url, exc)
        return None


def _generic_links(soup: BeautifulSoup, base_url: str, selector: str) -> list[dict]:
    items = []
    for tag in soup.select(selector):
        title = tag.get_text(strip=True)
        href  = tag.get("href", "")
        if not href or not title:
            continue
        if href.startswith("/"):
            p = urlparse(base_url)
            href = f"{p.scheme}://{p.netloc}{href}"
        elif not href.startswith("http"):
            continue
        items.append({"title": title, "url": href})
    return items


# Parser registry — one function per "parser" key in config
_PARSERS: dict[str, Callable] = {}


def _parser(name: str):
    def decorator(fn):
        _PARSERS[name] = fn
        return fn
    return decorator


@_parser("rbc")
def _rbc(soup, url): return _generic_links(soup, url, "h3 a, h2 a, .article-title a")

@_parser("td")
def _td(soup, url): return _generic_links(soup, url, "h3 a, h2 a, .publication-title a")

@_parser("scotiabank")
def _scotiabank(soup, url): return _generic_links(soup, url, "h3 a, h2 a, .pub-title a, a.pub-link")

@_parser("bmo")
def _bmo(soup, url): return _generic_links(soup, url, "h3 a, h2 a, article a")

@_parser("cibc")
def _cibc(soup, url): return _generic_links(soup, url, "h3 a, h2 a, .card-title a")

@_parser("nbc")
def _nbc(soup, url): return _generic_links(soup, url, "h3 a, h2 a, .article a")

@_parser("iea")
def _iea(soup, url): return _generic_links(soup, url, "h3 a, h2 a, .m-news-list__item a")

@_parser("spglobal")
def _spglobal(soup, url): return _generic_links(soup, url, "h3 a, h2 a, .cta-link, .article-title a")

@_parser("woodmac")
def _woodmac(soup, url): return _generic_links(soup, url, "h3 a, h2 a, .card__title a")

@_parser("mckinsey")
def _mckinsey(soup, url): return _generic_links(soup, url, "h3 a, h2 a, .article-link")

@_parser("deloitte")
def _deloitte(soup, url): return _generic_links(soup, url, "h3 a, h2 a, .article-title a")

@_parser("goldman")
def _goldman(soup, url): return _generic_links(soup, url, "h3 a, h2 a, .insight-card__title a")

@_parser("signal49")
def _signal49(soup, url): return _generic_links(soup, url, "h3 a, h2 a, article a")


def fetch_scrape(source: dict, cutoff: datetime, pattern: re.Pattern,
                 use_dedup: bool, seen_check: Callable, seen_mark: Callable) -> list[dict]:
    results = []
    parser_fn = _PARSERS.get(source.get("parser", ""))
    if not parser_fn:
        log.warning("  No parser for %s — skipping", source["name"])
        return results

    for url in source["urls"]:
        log.info("  SCRAPE %s → %s", source["name"], url)
        soup = _get_soup(url)
        if soup is None:
            continue

        links = parser_fn(soup, url)
        log.info("    %d candidate links", len(links))

        for item in links:
            article_url = item["url"]
            title       = item["title"]
            if use_dedup and seen_check(article_url):
                continue
            if not matches(pattern, title):
                continue
            results.append(_build_row(source, title, article_url, None, ""))
            if use_dedup:
                seen_mark(article_url)

        time.sleep(REQUEST_DELAY)
    return results


# ===========================================================================
# Main entry point (library + CLI)
# ===========================================================================

def run(
    sources: list[dict],
    keywords: list[str],
    days: int,
    run_id: str,
    use_dedup: bool = True,
    progress_callback: Optional[Callable[[str, int, int], None]] = None,
) -> list[dict]:
    """
    Scrape all sources and return a list of article dicts.

    Parameters
    ----------
    sources           : filtered list of source configs
    keywords          : keywords to match against title/summary
    days              : how many days back to look
    run_id            : identifier for this run (used for logging)
    use_dedup         : skip already-seen URLs (uses db.py helpers)
    progress_callback : called after each source as (source_name, done, total)
    """
    import db  # imported here to keep scraper.py usable without db.py

    pattern = build_pattern(keywords)
    cutoff  = datetime.now(timezone.utc) - timedelta(days=days)
    total   = len(sources)

    log.info("[%s] Starting scrape — %d sources, %d keywords, %d-day window",
             run_id, total, len(keywords), days)

    def seen_check(url): return db.is_seen(url) if use_dedup else False
    def seen_mark(url):
        if use_dedup:
            db.mark_seen(url)

    all_rows: list[dict] = []

    for i, source in enumerate(sources, start=1):
        try:
            if source["type"] == "rss":
                rows = fetch_rss(source, cutoff, pattern, use_dedup, seen_check, seen_mark)
            elif source["type"] == "scrape":
                rows = fetch_scrape(source, cutoff, pattern, use_dedup, seen_check, seen_mark)
            else:
                rows = []
            log.info("  → %d articles from %s", len(rows), source["name"])
            all_rows.extend(rows)
        except Exception as exc:
            log.error("FAILED %s: %s", source["name"], exc, exc_info=True)

        if progress_callback:
            progress_callback(source["name"], i, total)

    # Deduplicate by URL within this run
    seen_urls: set[str] = set()
    unique_rows = []
    for row in all_rows:
        u = row["URL"]
        if u not in seen_urls:
            seen_urls.add(u)
            unique_rows.append(row)

    log.info("[%s] Done — %d unique articles", run_id, len(unique_rows))
    return unique_rows


# ===========================================================================
# CLI entry point
# ===========================================================================

if __name__ == "__main__":
    import argparse
    import sys
    import uuid
    from pathlib import Path

    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill

    import config
    import db
    import settings as smod

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler("scraper.log", encoding="utf-8"),
        ],
    )

    p = argparse.ArgumentParser(description="NRCan Article Scraper (CLI)")
    p.add_argument("--days",     type=int, default=None)
    p.add_argument("--keywords", type=str, default="")
    p.add_argument("--output",   type=str, default="nrcan_articles")
    p.add_argument("--no-dedup", action="store_true")
    args = p.parse_args()

    db.init_db()
    cfg = smod.get_effective()

    days     = args.days or cfg["lookback_days"]
    keywords = cfg["keywords"] + (args.keywords.split() if args.keywords else [])
    sources  = cfg["sources"]
    rid      = uuid.uuid4().hex

    rows = run(
        sources=sources,
        keywords=keywords,
        days=days,
        run_id=rid,
        use_dedup=not args.no_dedup,
    )

    if not rows:
        print("No matching articles found.")
        sys.exit(0)

    db.insert_articles(rid, rows)

    out_dir = Path(config.OUTPUT_DIR)
    out_dir.mkdir(exist_ok=True)

    df = pd.DataFrame(rows).sort_values(["Category", "Source", "Published"],
                                        ascending=[True, True, False])
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    xlsx = out_dir / f"{args.output}_{ts}.xlsx"
    csv  = out_dir / f"{args.output}_{ts}.csv"

    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Articles")
        ws = writer.sheets["Articles"]
        for col, w in zip("ABCDEFG", [22, 18, 60, 50, 12, 80, 20]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        fill = PatternFill("solid", fgColor="1F4E79")
        font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.font      = font
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=4)
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")

    df.to_csv(csv, index=False, encoding="utf-8-sig")
    print(f"\n{len(df)} articles → {xlsx}")

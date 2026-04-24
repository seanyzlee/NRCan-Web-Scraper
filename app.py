"""
app.py — Flask web application for the NRCan Article Scraper.

Run locally:
    python app.py          → http://localhost:5000

Production (gunicorn):
    gunicorn app:app --workers 1 --threads 4 --bind 0.0.0.0:$PORT --timeout 120
"""

import io
import logging
import os
import threading
import time
import uuid
from datetime import datetime, timezone
from logging.handlers import RotatingFileHandler
from pathlib import Path

from flask import (Flask, flash, jsonify, redirect, render_template,
                   request, send_file, url_for)

from bs4 import BeautifulSoup

import db
import scraper
import settings as smod
import iran_config

# ---------------------------------------------------------------------------
# Paths  (DATA_DIR=/data on Render, cwd locally)
# ---------------------------------------------------------------------------
_DATA_DIR = Path(os.environ.get("DATA_DIR", "."))
_DATA_DIR.mkdir(parents=True, exist_ok=True)
(_DATA_DIR / "output").mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "nrcan-scraper-dev-key-change-in-prod")

# ---------------------------------------------------------------------------
# Logging  (file goes to DATA_DIR so it survives on the persistent disk)
# ---------------------------------------------------------------------------
_log_file = _DATA_DIR / "scraper.log"
_file_handler = RotatingFileHandler(str(_log_file), maxBytes=5_000_000,
                                    backupCount=3, encoding="utf-8")
_file_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
logging.getLogger().addHandler(_file_handler)
logging.getLogger().addHandler(logging.StreamHandler())   # also print to stdout
logging.getLogger().setLevel(logging.INFO)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# DB init — runs at import time so gunicorn picks it up automatically
# ---------------------------------------------------------------------------
db.init_db()
log.info("DB initialised at %s", db.DB_PATH)

# ---------------------------------------------------------------------------
# Scrape job state (module-level, single-worker gunicorn only)
# ---------------------------------------------------------------------------
_lock        = threading.Lock()
_running     = False
_current_run = None


def _is_running() -> bool:
    with _lock:
        return _running


def _start_running(run_id: str) -> None:
    global _running, _current_run
    with _lock:
        _running     = True
        _current_run = run_id


def _stop_running() -> None:
    global _running, _current_run
    with _lock:
        _running     = False
        _current_run = None


# ---------------------------------------------------------------------------
# Background scrape thread
# ---------------------------------------------------------------------------

def _scrape_thread(run_id: str, use_dedup: bool) -> None:
    cfg = smod.get_effective()

    def on_progress(source_name, done, total):
        db.update_run_progress(run_id, done)

    try:
        rows = scraper.run(
            sources=cfg["sources"],
            keywords=cfg["keywords"],
            days=cfg["lookback_days"],
            run_id=run_id,
            use_dedup=use_dedup,
            progress_callback=on_progress,
        )
        count = db.insert_articles(run_id, rows)
        db.finish_run(run_id, count)
        log.info("Run %s finished — %d articles", run_id, count)
    except Exception as exc:
        log.error("Run %s failed: %s", run_id, exc, exc_info=True)
        db.fail_run(run_id, str(exc))
    finally:
        _stop_running()


# ---------------------------------------------------------------------------
# Routes — pages
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    run_id   = request.args.get("run_id") or db.get_latest_run_id()
    all_runs = db.get_runs(limit=10)

    categories = []
    sources    = []
    articles   = []
    run_info   = None

    if run_id:
        run_info   = db.get_run(run_id)
        categories = db.get_distinct_values(run_id, "category")
        sources    = db.get_distinct_values(run_id, "source")
        articles   = db.get_articles(run_id)

    return render_template(
        "index.html",
        articles=articles,
        categories=categories,
        sources=sources,
        run_id=run_id,
        run_info=run_info,
        all_runs=all_runs,
        is_running=_is_running(),
        current_run=_current_run,
    )


@app.route("/settings", methods=["GET"])
def settings_page():
    cfg = smod.get_effective()
    return render_template("settings.html", cfg=cfg)


@app.route("/settings", methods=["POST"])
def settings_save():
    try:
        smod.save_from_form(request.form)
        flash("Settings saved.", "success")
    except Exception as exc:
        flash(f"Error saving settings: {exc}", "danger")
    return redirect(url_for("settings_page"))


# ---------------------------------------------------------------------------
# Routes — API
# ---------------------------------------------------------------------------

@app.route("/api/run", methods=["POST"])
def api_start_run():
    if _is_running():
        return jsonify({"error": "A scrape is already running."}), 409

    cfg = smod.get_effective()
    if not cfg["sources"]:
        return jsonify({"error": "No sources enabled. Check settings."}), 400
    if not cfg["keywords"]:
        return jsonify({"error": "No keywords configured. Check settings."}), 400

    use_dedup = request.json.get("dedup", True) if request.is_json else True

    run_id = uuid.uuid4().hex
    db.create_run(run_id, total_sources=len(cfg["sources"]))
    _start_running(run_id)

    t = threading.Thread(target=_scrape_thread, args=(run_id, use_dedup), daemon=True)
    t.start()

    return jsonify({"run_id": run_id})


@app.route("/api/run/<run_id>")
def api_run_status(run_id: str):
    info = db.get_run(run_id)
    if not info:
        return jsonify({"error": "Run not found"}), 404
    return jsonify(info)


@app.route("/api/articles")
def api_articles():
    run_id   = request.args.get("run_id") or db.get_latest_run_id()
    category = request.args.get("category", "")
    source   = request.args.get("source", "")
    q        = request.args.get("q", "")

    if not run_id:
        return jsonify({"run_id": None, "articles": [], "total": 0})

    articles = db.get_articles(run_id, category=category, source=source, q=q)
    return jsonify({"run_id": run_id, "articles": articles, "total": len(articles)})


@app.route("/api/runs")
def api_runs():
    return jsonify(db.get_runs())


@app.route("/download/<run_id>")
def download_run(run_id: str):
    """Stream an Excel file for a completed run."""
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill

    rows = db.get_articles(run_id)
    if not rows:
        flash("No articles found for this run.", "warning")
        return redirect(url_for("index"))

    # rows are plain dicts; rename keys for the spreadsheet header
    rename = {
        "source": "Source", "category": "Category", "title": "Title",
        "url": "URL", "published": "Published", "summary": "Summary",
        "scraped_at": "Scraped At",
    }
    df = pd.DataFrame(rows).rename(columns=rename)
    df = df[["Source", "Category", "Title", "URL", "Published", "Summary", "Scraped At"]]
    df = df.sort_values(["Category", "Source", "Published"], ascending=[True, True, False])

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Articles")
        ws = writer.sheets["Articles"]
        for col, w in zip("ABCDEFG", [22, 18, 60, 50, 12, 80, 20]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        fill     = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.font      = hdr_font
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for ri in range(2, ws.max_row + 1):
            cell = ws.cell(row=ri, column=4)
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")

    buf.seek(0)
    filename = f"nrcan_articles_{run_id[:8]}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------------------
# Iran Monitor — prices cache
# ---------------------------------------------------------------------------

_prices_cache: dict = {}
_prices_lock  = threading.Lock()
_PRICES_TTL   = 300   # seconds (5 min)

# Serialises all yfinance calls — prevents concurrent SQLite cache writes
# (which cause "database is locked") and avoids Yahoo rate-limiting from
# parallel requests.
_yf_lock = threading.Lock()

# Iran news cache
_iran_news_cache: dict = {"articles": [], "fetched_at": 0}
_iran_news_lock  = threading.Lock()
_IRAN_NEWS_TTL   = 1800  # 30 min


@app.route("/iran")
def iran():
    return render_template("iran.html", iran_cfg=iran_config)


@app.route("/api/iran/prices")
def api_iran_prices():
    """Return live commodity + stock prices via yfinance (cached 5 min).

    Fetches each ticker sequentially inside _yf_lock to avoid:
    - Yahoo Finance rate-limiting (caused by yf.download's internal thread pool)
    - SQLite 'database is locked' on yfinance's timezone cache
    """
    now = time.time()
    with _prices_lock:
        if _prices_cache and now - _prices_cache.get("_ts", 0) < _PRICES_TTL:
            return jsonify(_prices_cache)

    try:
        import yfinance as yf

        all_meta = (
            [(c, "commodity") for c in iran_config.COMMODITIES]
            + [(s, "stock")     for s in iran_config.STOCKS]
        )

        raw: dict[str, tuple[float, float]] = {}

        with _yf_lock:
            for meta, _ in all_meta:
                symbol = meta["symbol"]
                try:
                    fi    = yf.Ticker(symbol).fast_info
                    price = float(fi.last_price     or 0)
                    prev  = float(fi.previous_close or price)
                    raw[symbol] = (price, prev)
                except Exception as exc:
                    log.warning("Price fetch skipped %s: %s", symbol, exc)
                    raw[symbol] = (0.0, 0.0)
                time.sleep(0.3)   # stay well under Yahoo's rate limit

        def _fmt(meta: dict) -> dict:
            symbol = meta["symbol"]
            price, prev = raw.get(symbol, (0.0, 0.0))
            change     = price - prev
            change_pct = (change / prev * 100) if prev else 0.0
            return {
                "symbol":     symbol,
                "label":      meta.get("label", symbol),
                "price":      round(price, 2),
                "change":     round(change, 2),
                "change_pct": round(change_pct, 2),
                "unit":       meta.get("unit", "USD"),
                "exchange":   meta.get("exchange", ""),
            }

        result = {
            "commodities": [_fmt(c) for c in iran_config.COMMODITIES],
            "stocks":      [_fmt(s) for s in iran_config.STOCKS],
            "updated_at":  datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
            "_ts":         now,
        }

        with _prices_lock:
            _prices_cache.clear()
            _prices_cache.update(result)

        return jsonify(result)

    except Exception as exc:
        log.error("Iran prices fetch failed: %s", exc, exc_info=True)
        return jsonify({"error": str(exc), "commodities": [], "stocks": []}), 500


@app.route("/api/iran/news")
def api_iran_news():
    """Return Iran-filtered news (cached 30 min).

    Each RSS source is fetched in parallel with an 8-second timeout so a
    single slow/dead feed cannot block the entire response.
    """
    now   = time.time()
    force = request.args.get("force") == "1"
    with _iran_news_lock:
        if not force and _iran_news_cache["articles"] and \
                now - _iran_news_cache["fetched_at"] < _IRAN_NEWS_TTL:
            return jsonify({
                "articles":   _iran_news_cache["articles"],
                "cached":     True,
                "fetched_at": _iran_news_cache["fetched_at"],
            })

    import feedparser
    from concurrent.futures import ThreadPoolExecutor, as_completed

    pattern = scraper.build_pattern(iran_config.KEYWORDS)

    _headers = {
        "User-Agent": "Mozilla/5.0 (compatible; NRCanBot/1.0)",
        "Accept": "application/rss+xml, application/xml, text/xml, */*",
    }

    def _classify_topics(title: str, summary: str) -> list[str]:
        text = (title + " " + summary).lower()
        matched = [
            topic for topic, kws in iran_config.TOPIC_KEYWORDS.items()
            if any(kw in text for kw in kws)
        ]
        return matched or ["General"]

    def _fetch_url(source_name: str, url: str, pre_filtered: bool) -> list[dict]:
        """Fetch and parse one RSS URL. One future per URL keeps timeouts predictable."""
        found = []
        try:
            resp = scraper.SESSION.get(url, timeout=10, headers=_headers)
            resp.raise_for_status()
            feed = feedparser.parse(resp.content)
        except Exception as exc:
            log.warning("Iran RSS skip (%s): %s", url, exc)
            return found

        for entry in feed.entries:
            title = entry.get("title", "").strip()
            link  = entry.get("link",  "").strip()
            if not title or not link:
                continue
            if not pre_filtered and not scraper.matches(pattern, title):
                continue

            pub_raw = entry.get("published_parsed") or entry.get("updated_parsed")
            pub_dt  = scraper.parse_date(pub_raw)

            summary_raw = entry.get("summary", "") or entry.get("description", "")
            summary = BeautifulSoup(summary_raw, "lxml").get_text(" ", strip=True)[:300] \
                      if summary_raw else ""

            found.append({
                "source":    source_name,
                "title":     title,
                "url":       link,
                "published": pub_dt.strftime("%Y-%m-%d") if pub_dt else "",
                "summary":   summary,
                "topics":    _classify_topics(title, summary),
            })
        return found

    # One future per URL so a slow source never blocks others beyond its own timeout
    articles: list[dict] = []
    with ThreadPoolExecutor(max_workers=6) as pool:
        futures = {
            pool.submit(_fetch_url, src["name"], url, src.get("pre_filtered", False)): url
            for src in iran_config.SOURCES
            for url in src["urls"]
        }
        try:
            for future in as_completed(futures, timeout=30):
                try:
                    articles.extend(future.result())
                except Exception as exc:
                    log.warning("Iran news future error: %s", exc)
        except TimeoutError:
            log.warning("Iran news: some URLs did not complete within 30s — returning partial results")
            for future in futures:
                if future.done():
                    try:
                        articles.extend(future.result())
                    except Exception:
                        pass

    # Deduplicate by URL, sort newest first
    seen_urls: set[str] = set()
    unique = []
    for a in sorted(articles, key=lambda x: x["published"], reverse=True):
        if a["url"] not in seen_urls:
            seen_urls.add(a["url"])
            unique.append(a)

    with _iran_news_lock:
        _iran_news_cache["articles"]   = unique
        _iran_news_cache["fetched_at"] = now

    return jsonify({"articles": unique, "cached": False, "fetched_at": now})


# ---------------------------------------------------------------------------
# Local dev entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print(f"NRCan Article Scraper → http://localhost:5000  (data: {_DATA_DIR})")
    app.run(debug=False, host="0.0.0.0", port=5000)
